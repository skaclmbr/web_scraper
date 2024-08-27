# # pulls down data from eBird website,
# # populates MongoDB implementation,
# # feeds the Block Explorer

import requests
from bs4 import BeautifulSoup
import json
from pymongo.mongo_client import MongoClient
# import mdbconn #stores database connection information
import certifi
from datetime import datetime
import os

from openpyxl import Workbook #allows connecting to databases
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle

#############################################################################
## LOOKUP DICTS

obs_status = {
    "Observed" : "C1",
    "Possible" : "C2",
    "Probable" : "C3",
    "Confirmed" : "C4"
}

nl =" \n"
fmt_dt = "%Y-%m-%d"
today = datetime.now().strftime(fmt_dt)

#############################################################################
## OUTPUT EXCEL FILE
dirpath = os.path.dirname(os.path.abspath(__file__))
wb = Workbook()

#field formats
pctFields = [
    "breedPctPossible",
    "breedPctConfirmed",
    "breedPctProbable"
]
bFields = [
    "bbcgCoded",
    "bbcgConfirmed",
    "bbcgPossible",
    "bbcgTotalEffortHours"
]

# Stats Sheet
sc = {
    "A" : "ID_EBD_NAME",
    "B" : "tot",
    "C" : "con",
    "D" : "breedPctConfirmed",
    "E" : "pos",
    "F" : "breedPctPossible",
    "G" : "pro",
    "H" : "breedPctProbable",
    "I" : "diurnal_hrs",
    "J" : "nocturnal_hrs",
    "K" : "num_checklists",
    "L" : "num_atlasers",
    "M" : "status",
    "N" : "REGION",
    "O" : "ECOREGION",
    "P" : "COUNTY",
    "Q" : "updateDate",
    "R" : "ID_BLOCK_CODE",
    "S" : "ID_NCBA_BLOCK",
    "T" : "bbcgCoded",
    "U" : "bbcgConfirmed",
    "V" : "bbcgPossible",
    "W" : "bbcgTotalEffortHrs"
}
colsStats = {}
wsStats = wb.active
wsStats.title = "Stats"
maxStatCol = ""

for k,v in sc.items():
    wsStats[ k + "1"] = v
    # flip these for lookup while saving data
    colsStats[v] = k
    maxStatCol = k


#species sheet
sppc = {
    "A" : "ID_EBD_NAME",
    "B" : "common_name",
    "C" : "breeding_evidence",
    "D" : "breeding_code",
    "E" : "breeding_category",
    "F" : "recent_location",
    "G" : "observation_date",
    "H" : "sampling_event_identifier",
    "I" : "status",
    "J" : "REGION",
    "K" : "ECOREGION",
    "L" : "COUNTY",
    "M" : "updateDate",
    "N" : "ID_BLOCK_CODE",
    "O" : "ID_NCBA_BLOCK"
}
colsSpp = {}
wsSpp = wb.create_sheet("Species")
maxSppCol = ""
for k,v in sppc.items():
    wsSpp[ k + "1"] = v
    colsSpp[v] = k
    maxSppCol = k

# Top Atlasers sheet
tac = {
    "A" : "ID_EBD_NAME",
    "B" : "rank",
    "C" : "observer",
    "D" : "confirmed-species",
    "E" : "status",
    "F" : "REGION",
    "G" : "ECOREGION",
    "H" : "COUNTY",
    "I" : "updateDate",
    "J" : "ID_BLOCK_CODE",
    "K" : "ID_NCBA_BLOCK"
}
colsTA = {}
wsTopAtlasers = wb.create_sheet("Top Atlasers")
maxTACol = ""

for k,v in tac.items():
    wsTopAtlasers[ k + "1"] = v
    colsTA[v] = k
    maxTACol = k

# Recent Visits sheet
rvc = {
    "A" : "ID_EBD_NAME",
    "B" : "observer",
    "C" : "observation_date",
    "D" : "checklist",
    "E" : "observation_start_time",
    "F" : "recent_location",
    "G" : "num_spp",
    "H" : "REGION",
    "I" : "ECOREGION",
    "J" : "COUNTY",
    "K" : "updateDate",
    "L" : "ID_BLOCK_CODE",
    "M" : "ID_NCBA_BLOCK"
}    
colsRV = {}
wsRecentVisits = wb.create_sheet("Recent Visits")
maxRVCol = ""
for k,v in rvc.items():
    wsRecentVisits[ k + "1"] = v
    colsRV[v] = k
    maxRVCol = k


#############################################################################
## EXCEL FUNCTIONS
style = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
    )

red_fill = PatternFill(bgColor="FF7979")
green_fill = PatternFill(bgColor="92D050")
pct_style = "0.0%"

# dxRed = DifferentialStyle(font=Font(bold=True), fill=red_fill)
# dxGreen = DifferentialStyle(fill=green_fill)

# rulConRed = 

# def addCondFormat(ws, range, formula):
#     pass


def createTable(dname, maxCol, currRow, style = style):
    tbl = Table(
        displayName = dname,
        ref = "A1:" + maxCol + str(currRow)
    )
    tbl.tableStyleInfo = style
    return tbl

#############################################################################
## DATE FUNCTIONS
def getJDay (d):
    dt = datetime.strptime(d, fmt_dt)
    dtt = dt.timetuple()
    return dtt.tm_yday

breeding_start = getJDay("2021-03-01")
breeding_end = getJDay("2021-08-15")
breeding_length = breeding_end - breeding_start
winter_start = getJDay("2021-11-01")
winter_end = getJDay("2022-03-01")

winter1_end = getJDay("2021-12-31")
winter2_start = winter1_end + 1

breeding1_end = breeding_start + (breeding_length/3)
breeding2_start = breeding1_end + 1
breeding2_end = breeding2_start + (breeding_length/3)
breeding3_start = breeding2_end + 1

update_date = today

def fmt_date(t):

    dt = datetime.strptime(t, "%d %b %Y" )

    return dt.strftime("%Y-%m-%d")

#############################################################################
## block completion criteria
minBlockBreedDiurnalHrs = 20
minBlockWinterDiurnalHrs = 5
minBlockBreedVisits = 3
minBlockWinterVisits = 2
minBlockBreedNocturnalVisits = 3
minBlockWinterNocturnalVisits = 1
minBlockWinterDetected = 55
minBlockBreedCoded = 55
minBlockBreedConfirmedPct = 0.25
maxBlockBreedPossiblePct = 0.25

#############################################################################
## SETUP CONNECTION TO MONGODB
connString = (
    "mongodb+srv://ncba_admin:hydroprognecaspia@" + 
    "cluster0.rzpx8.mongodb.net/ebd_mgmt?retryWrites=true&w=majority"
    )

# connstring = "mongodb+srv://ncba_admin:hydroprognecaspia@cluster0.rzpx8.mongodb.net/ebd_mgmt?retryWrites=true&w=majority"
client = MongoClient(connString, tlsCAFile=certifi.where())

db = client.ebd_mgmt
blocks = db.blocks
ewd = db.ebird_web_data
blocksum = db.BLOCK_SUMMARIES


# get list of priority blocks

cursor = blocks.find(
    {
        "PRIORITY" : "1"
    },
    {
        "_id" : 1,
        "ID_EBD_NAME" : 1,
        "ID_NCBA_BLOCK" : 1,
        "ECOREGION" : 1,
        "REGION" : 1,
        "COUNTY" : 1,
        "ID_BLOCK_CODE" : 1
    }
)

pblocks = {}
for b in list(cursor):
    pblocks[b["ID_BLOCK_CODE"]] = b


def get_species(tbody):

    results = []

    for r in tbody.findChildren("tr"):
        bird = {}
        # common name
        bird["common_name"] = r.find(
            "td", 
            attrs={"headers": "th-species-name"}
            ).text
        
        # breeding code
        obs = r.find(
            "td", 
            attrs={"headers": "th-summary-obs"}
            ).text
        
        if "(" in obs:

            obs = obs.split("(")
            bird["breeding_evidence"] = obs[0].strip()
            bird["breeding_code"] = obs[1].replace(")","").strip()
            try:
                bird["breeding_category"] = obs_status[bird["breeding_evidence"]]
            except:
                bird["breeding_category"] = ""

        else:
            bird["breeding_evidence"] = obs.strip()
            bird["breeding_category"] = "C1"

        # location
        bird["recent_location"] = r.find(
            "td",
            attrs={"headers" : "th-summary-loc"}
        ).text

        # date/checklist
        recent_dt = r.find(
            "td", 
            attrs={"headers": "th-summary-date"}
            )
        bird["observation_date"] = fmt_date(recent_dt.text)
        bird["sampling_event_identifier"] = recent_dt.find("a")["href"].split("/")[-1]
    
        results.append(bird)

    return results



def parse_page(parsed_html):
    
    bd = {
        "updateDate" : update_date,
        "breedPctConfirmed" : 0 ,
        "breedPctProbable" : 0 ,
        "breedPctPossible" : 0 ,
        "bbcgCoded" : 0,
        "bbcgConfirmed" : 0,
        "bbcgPossible" : 0,
        "bbcgTotalEffortHrs" : 0
        }
    #############################################################################
    # GET BLOCK INFO
    name_header = parsed_html.find(
        "h1",
        attrs={
            "class" : "hotspot--name"
            }
        )
    for child in name_header.find_all("a"):
        child.decompose()

    bd["ebird_block_name"] = name_header.text.strip()

    #############################################################################
    # GET BLOCK STATUS
    for bd_headers in parsed_html.find_all("div", attrs={"class": "bd"}):
        item = bd_headers.p.text

        match item:
            case "Status:":
                t = bd_headers.find_next("h2").text
                t = t.split(" ")
                bd["status"] = t[0] 
            case "Block type:":
                if "Priority" in bd_headers.find_next("h2").text:
                    bd["priority"] = 1
                else:
                    bd["priority"] = 0
            case "Effort hours (diurnal/nocturnal):":
                temp = bd_headers.find_next("h2").text
                temp = temp.split("/")

                bd["diurnal_hrs"] = float(temp[0].replace(",","").strip())
                bd["nocturnal_hrs"] = float(temp[1].replace(",","").strip())

            case _:
                pass
            
    if (
        bd["diurnal_hrs"] >= minBlockBreedDiurnalHrs
        ):
        bd["bbcgTotalEffortHrs"] =  1
    else: bd["bbcgTotalEffortHrs"] = 0
            
    #############################################################################
    # GET CHECKLISTS/ATLASERS
    for s in parsed_html.find_all("span", attrs={"class": None}):
        match s.text.strip():
            case "Checklists":
                bd["num_checklists"] = int(s.find_previous("span").text)
            case "Atlasers":
                bd["num_atlasers"] = int(s.find_previous("span").text)


    #############################################################################
    # GET OBS/POSS/PROB/CONF/TOTAL
    tr = parsed_html.find("tr", attrs={"class": "tr--major"})

    for d in tr.findChildren("td"):

        h = d["headers"][0]
        h = h.split("-")[2]
        if h != "period":
            bd[h] = int(d.text)

    if (
        bd["tot"] >= minBlockBreedCoded
    ):
        bd["bbcgCoded"] = 1
    else: bd["bbcgCoded"] = 0
    
    try:
        bd["breedPctConfirmed"] = bd["con"]/bd["tot"]
        if ( 
            bd["breedPctConfirmed"] > minBlockBreedConfirmedPct
            ):
            bd["bbcgConfirmed"] = 1
        else: 0
    except: pass
    try:  
        bd["breedPctProbable"] = bd["pro"]/bd["tot"]
    except: pass

    try:
        bd["breedPctPossible"] = bd["pos"]/bd["tot"]
        if( 
            bd["breedPctPossible"] <= maxBlockBreedPossiblePct
            ):
            bd["bbcgPossible"] = 1
        else: bd["bbcgPossible"] = 0
    except:
        pass
    #############################################################################
    # GET SPP DATA
    try:
        spp_head = parsed_html.find("h3", attrs={"class": "mbm"})
        tbody = spp_head.find_next("tbody")
        bd["spp"] = get_species(tbody)
    except:
        bd["spp"] = []

    #############################################################################
    # GET TOP ATLASERS
    bd["top_atlasers"] = []
    try:
        ta_table = parsed_html.find("table", attrs={"id" : "top-ebirders"})

        for r in ta_table.find_all("tr"):
            ta = {}
            #top-rank
            ta["rank"] = int(r.find(
                "td",
                attrs={
                    "headers": "rank-top-ebirders"}
                ).text)
            
            #observer
            ta["observer"] = r.find(
                "td",
                attrs = {
                    "headers" : "observer"
                }
            ).text

            ta["confirmed-species"] = int(r.find(
                "div",
                attrs = {
                    "class" : "hist-bar--percent"
                }
            ).text)
            
            bd["top_atlasers"].append(ta)
    except:
        pass

    return bd


def parse_recent_visits(html):
    rv = [] # list of recent visits
    try:
        table = html.find(
            "table",
            attrs = {
                "class" : "table"
            }
        )
        tbody = table.find("tbody")
        for r in tbody.find_all(
            "tr"
        ):
            rvd = {} #recent visit dict
            # observer
            rvd["observer"] = r.find("td", attrs={"headers" : "observer"}).text
            # date/checklist
            col2 = r.find("td", attrs = {"headers" : "date"}).find("a")
            rvd["observation_date"] = fmt_date(col2.text.strip())
            rvd["checklist"] = col2["href"].split("/")[-1]

            # obs time
            rvd["observation_start_time"] = r.find(
                "td",
                attrs={"class" : "obstable-time"}
                ).text.strip()
            
            # recent location
            rvd["recent_location"] = r.find(
                "td",
                attrs={"class" : "obstable-location"}
                ).text.strip()

            # num spp
            rvd["num_spp"] = int(
                r.find(
                    "td",
                    attrs={"class" : "obstable-species"}
                    ).text
                )

            rv.append(rvd)
    except:
        pass

    return rv

def main():
    # loop through blocks, hit eBird webpage, and parse data

    # Excel Row counters
    # row counters
    rowST = 2
    rowTA = 2
    rowSP = 2
    rowRV = 2

    # BLOCK COUNTER
    bcount = 0
    for b in pblocks.keys():
        ## add basic block data
        pd = pblocks[b]

        print(
            nl,
            "=================", nl,
            "retrieving eBird page for", b
        )

        # get html from overview web page
        r= requests.get("https://ebird.org/atlasnc/block/" + b)

        # parse retrieved html
        full_page_html = BeautifulSoup(r.content, "html.parser")
        content_html = full_page_html.find("div", attrs = {"class": "page"}) 
        pd.update(parse_page(content_html))

        # get html for recent visits page
        r= requests.get("https://ebird.org/atlasnc/block/" + b + "/activity")
        full_page_html = BeautifulSoup(r.content, "html.parser")
        content_html = full_page_html.find("div", attrs = {"class": "page"}) 
        prv = parse_recent_visits(content_html)

        pd["recent_visits"] = prv

        print("data parsed for block", b)
        print("saving data to excel")

        # POPULATE EXCEL

        for k,v in pd.items():
            if k == "recent_visits":
                for j in v:
                    # loop through items in array
                    for x, y in j.items():
                        #loop through fields
                        if x == "checklist":
                            wsRecentVisits[colsRV[x]+str(rowRV)] = (
                                '=HYPERLINK("https://ebird.org/atlasnc/checklist/' +
                                y + '", "' + y + '")'
                            )
                            
                        else:
                            wsRecentVisits[colsRV[x]+str(rowRV)] = y

                    
                    #add stats cols
                    for x,y in colsStats.items():
                        if x in colsRV.keys():
                            wsRecentVisits[colsRV[x] + str(rowRV)] = pd[x]

                    rowRV += 1
            elif k == "spp":
                for j in v:
                    # loop through items in array
                    for x, y in j.items():
                        #loop through fields
                        if x == "sampling_event_identifier":
                            wsSpp[colsSpp[x]+str(rowSP)] = (
                                '=HYPERLINK("https://ebird.org/atlasnc/checklist/' +
                                y + '", "' + y + '")'
                            )
                        else:
                            wsSpp[colsSpp[x]+str(rowSP)] = y
                                            
                    #add stats cols
                    for x,y in colsStats.items():
                        if x in colsSpp.keys():
                            wsSpp[colsSpp[x] + str(rowSP)] = pd[x]

                    rowSP += 1
            elif k == "top_atlasers":
                for j in v:
                    # loop through items in array
                    for x, y in j.items():
                        #loop through fields

                        wsTopAtlasers[colsTA[x]+str(rowTA)] = y
                       
                    #add stats cols
                    for x,y in colsStats.items():
                        if x in colsTA.keys():
                            wsTopAtlasers[colsTA[x] + str(rowTA)] = pd[x]

                    rowTA += 1
            else: #all other fields
                try:
                    wsStats[colsStats[k] + str(rowST)] = v
                except: pass

                if k in pctFields:
                    _c = wsStats[colsStats[k]+str(rowST)]
                    _c.number_format = "0.0%"
                #if field in another sheet, add



        #advance row for next block
        rowST += 1              
              
        print(nl,
            "uploading to atlas cache"
            )
        
        # update mongodb ebird data

        # blocksum.update_one(
        #     {
        #         "ID_BLOCK_CODE" : b
        #     },
        #     {
        #         "$set": {
        #             "ebird_web_data": pd
        #         }
        #     }
        # )

        bcount += 1
        print("atlas cache updated...")
        print(str(bcount), "blocks updated.", nl)

        # # TESTING
        # # if bcount > 19: break

    # SAVE WORKBOOK
    # Add tables

    # backup row counters
    rowRV -= 1
    rowSP -= 1
    rowST -= 1
    rowTA -= 1

    wsStats.add_table(
        createTable("blockstats", maxStatCol, rowST)
    )

    ## add conditional formatting
    ## Pct Confirmed
    pctConfCol = colsStats["breedPctConfirmed"]
    condCell1 = pctConfCol + "2"
    condRng = pctConfCol + "2:" + pctConfCol + str(rowST) 
    wsStats.conditional_formatting.add(
        condRng,
        FormulaRule(
            formula = [
                condCell1 + '<' + str(minBlockBreedConfirmedPct)
            ],
            stopIfTrue = False,
            fill = red_fill
        )
    )

    wsStats.conditional_formatting.add(
        condRng,
        FormulaRule(
            formula = [
                condCell1 + ">=" + str(minBlockBreedConfirmedPct)
            ],
            stopIfTrue = True,
            fill= green_fill
        )
    )
    ## Pct Possible
    pctPossCol = colsStats["breedPctPossible"] 
    condCell1 = pctPossCol + "2"
    condRng = pctPossCol + "2:" + pctPossCol + str(rowST) 
    wsStats.conditional_formatting.add(
        condRng,
        FormulaRule(
            formula = [
                condCell1 + '<' + str(maxBlockBreedPossiblePct)
            ],
            stopIfTrue = False,
            fill = green_fill
        )
    )

    wsStats.conditional_formatting.add(
        condRng,
        FormulaRule(
            formula = [
                condCell1 + ">=" + str(maxBlockBreedPossiblePct)
            ],
            stopIfTrue = True,
            fill= red_fill
        )
    )
    
    ## Coded
    condCol = colsStats["tot"] 
    condCell1 = condCol + "2"
    condRng = condCol + "2:" + condCol + str(rowST) 
    wsStats.conditional_formatting.add(
        condRng,
        FormulaRule(
            formula = [
                condCell1 + '>=' + str(minBlockBreedCoded)
            ],
            stopIfTrue = False,
            fill = green_fill
        )
    )

    wsStats.conditional_formatting.add(
        condRng,
        FormulaRule(
            formula = [
                condCell1 + "<" + str(minBlockBreedCoded)
            ],
            stopIfTrue = True,
            fill= red_fill
        )
    )
    
    ## Diurnal Hours
    condCol = colsStats["diurnal_hrs"] 
    condCell1 = condCol + "2"
    condRng = condCol + "2:" + condCol + str(rowST) 
    wsStats.conditional_formatting.add(
        condRng,
        FormulaRule(
            formula = [
                condCell1 + '>=' + str(minBlockBreedDiurnalHrs)
            ],
            stopIfTrue = False,
            fill = green_fill
        )
    )

    wsStats.conditional_formatting.add(
        condRng,
        FormulaRule(
            formula = [
                condCell1 + "<" + str(minBlockBreedDiurnalHrs)
            ],
            stopIfTrue = True,
            fill= red_fill
        )
    )

    wsSpp.add_table(
        createTable("species", maxSppCol, rowSP)
    )
    wsTopAtlasers.add_table(
        createTable("topatlasers", maxTACol, rowTA)
    )
    wsRecentVisits.add_table(
        createTable("recentvisits", maxRVCol, rowRV)
    )

    wb.save(
        os.path.join(
            dirpath,
            update_date + " ebird block stats.xlsx"
        )
    )    

if __name__=="__main__":
    main();