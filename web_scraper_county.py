# # pulls down data from eBird website,
# # populates MongoDB implementation,
# # feeds the Block Explorer

import requests
from bs4 import BeautifulSoup
import json
from pymongo.mongo_client import MongoClient
# import mdbconn #stores database connection information
import certifi
from datetime import datetime
import os
import csv
# import pandas as pd
import numpy as np
import time
from urllib.request import urlretrieve
from mdb_config import mdblogin_user, mdblogin_pass #contains login info for Mongo DB implementation

from openpyxl import Workbook #allows connecting to databases
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle

#############################################################################
## LOOKUP DICTS

nl =" \n"
fmt_dt = "%Y-%m-%d"
today = datetime.now().strftime(fmt_dt)
outdelim = ","
months = [
    "blank", # aligns key values with month values
    "Jan",
    "Feb",
    "Mar",
    "Apr",
    "May",
    "Jun",
    "Jul",
    "Aug",
    "Sep",
    "Oct",
    "Nov",
    "Dec"
]

csl = {}
scd = {}

try:
    with open("sppdata.json","r") as file:
        scd = json.load(file)
except:
    scd = {}
try:
    with open("spplist.json","r") as file:
        csl = json.load(file)
except:
    csl = {}


#############################################################################
## RETRIEVE COUNTY LIST

counties = {} # key = name, value = subnat2 code
with open("ebird_region_codes_nc.csv", mode = 'r') as file:
    csvFile = csv.reader(file)
    lcount = 0
    for lines in csvFile:
        if lcount != 0:
            counties[lines[5]] = lines[4]

        lcount += 1

#############################################################################
## OUTPUT EXCEL FILE
dirpath = os.path.dirname(os.path.abspath(__file__))
wb = Workbook()

#field formats
# # pctFields = [
# #     "breedPctPossible",
# #     "breedPctConfirmed",
# #     "breedPctProbable"
# # ]
# # bFields = [
# #     "bbcgCoded",
# #     "bbcgConfirmed",
# #     "bbcgPossible",
# #     "bbcgTotalEffortHours"
# # ]

# #species county sheet
sppc = {
    "A" : "County",
    "B" : "SpeciesName",
    "C" : "SpeciesCode",
    "D" : "SpeciesCategory",
    "E" : "SpeciesType",
    "F" : "SpeciesGraphUrl",
    "G" : "Phenology",
    "H" : "BreedAbundance",
    "I" : "MigAbundance",
    "J" : "WinterAbundance",
    "K" : "YearAbundance",
    "L" : "WeeksDetected"
}
colsSpp = {}
wsSpp = wb.active
wsSpp.title = "Species"
maxSppCol = ""
for k,v in sppc.items():
    wsSpp[ k + "1"] = v
    colsSpp[v] = k
    maxSppCol = k


# Data Sheet
# # sc = {
# #     "A" : "County",
# #     "B" : "SpeciesName",
# #     "C" : "SpeciesCode",
# #     "D" : "SpeciesCategory",
# #     "E" : "SpeciesType",
# #     "F" : "SpeciesGraphUrl",
# #     "G" : "Month",
# #     "H" : "Week",
# #     "I" : "BarSize",
# #     "J" : "BarSizeNum"
# #  }
# # colsData = {}
# # wsData = wb.create_sheet("Data")
# # maxDataCol = ""

# # for k,v in sc.items():
# #     wsData[ k + "1"] = v
# #     # flip these for lookup while saving data
# #     colsData[v] = k
# #     maxDataCol = k

#############################################################################
## EXCEL FUNCTIONS
style = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
    )

red_fill = PatternFill(bgColor="FF7979")
green_fill = PatternFill(bgColor="92D050")
pct_style = "0.0%"

def createTable(dname, maxCol, currRow, style = style):
    tbl = Table(
        displayName = dname,
        ref = "A1:" + maxCol + str(currRow)
    )
    tbl.tableStyleInfo = style
    return tbl

#############################################################################
## DATE FUNCTIONS
def getJDay (d):
    dt = datetime.strptime(d, fmt_dt)
    dtt = dt.timetuple()
    return dtt.tm_yday

breeding_start = getJDay("2021-03-01")
breeding_end = getJDay("2021-08-15")
breeding_length = breeding_end - breeding_start
winter_start = getJDay("2021-11-01")
winter_end = getJDay("2022-03-01")

winter1_end = getJDay("2021-12-31")
winter2_start = winter1_end + 1

breeding1_end = breeding_start + (breeding_length/3)
breeding2_start = breeding1_end + 1
breeding2_end = breeding2_start + (breeding_length/3)
breeding3_start = breeding2_end + 1

update_date = today

def fmt_date(t):

    dt = datetime.strptime(t, "%d %b %Y" )

    return dt.strftime("%Y-%m-%d")

#############################################################################
## SETUP CONNECTION TO MONGODB
connString = (
    "mongodb+srv://% s:% s@cluster0.rzpx8.mongodb.net/ebd_mgmt?retryWrites=true&w=majority" % (
        mdblogin_user,
        mdblogin_pass
    )
    )

client = MongoClient(connString, tlsCAFile=certifi.where())

db = client.ebd_mgmt
blocks = db.blocks
# ewd = db.ebird_web_data
blocksum = db.BLOCK_SUMMARIES

def calc_barsizenum(d):

    try: 
        response = int(d[-1:])
    except:
        response = 0

    return response


def get_data(table):
    # retrieve tab-separated file from web
    # return table json
    results = {}
    tbody = table.find("tbody")
    # loop through rows
    for r in tbody.findChildren("tr"):
        # loop through columns
        colnum = 1
        bird = {}
        bird_key = ""
        bird_data = []

        for c in r.find_all("td"):
            if not(c.find("th")):
                match colnum:

                    case 1: # SpeciesName
                        # SpeciesName, SpeciesCode, SpeciesCategory, SpeciesGraphUrl
                        spancount = 1
                        for i in c.find_all("span"):
                            # print(i.text)
                            match spancount:
                                case 1:
                                    link = i.find("a")
                                    if link:
                                        bird["SpeciesName"] = remove_chars(link.text)
                                        bird["SpeciesCode"] = link.get("data-species-code")                            
                                    else:
                                        bird["SpeciesName"] = remove_chars(i.text)
                                        bird["SpeciesCode"] = ""

                                    bird_key = bird["SpeciesName"]
                                    if " sp." in bird_key:
                                        bird["SpeciesCategory"] = "spuh"
                                    elif "/" in bird_key:
                                        bird["SpeciesCategory"] = "slash"
                                    elif "(hybrid)" in bird_key:
                                        bird["SpeciesCategory"] = "hybrid"
                                    else:
                                        bird["SpeciesCategory"] = "species"
                                case 2:
                                    if "(Domestic type)" in bird_key:
                                        bird["SpeciesType"] = "domestic"                        
                                    else:
                                        bird["SpeciesType"] = i.get("title")                        
                                case _:
                                    pass
                            spancount += 1
                        if "SpeciesType" not in bird.keys():
                            bird["SpeciesType"] = "native"
                        # print(bird_key)
                    case 2: # map
                        pass
                    case 3: # line graphs
                        bird["SpeciesGraphUrl"] = (
                            "https://ebird.org" + c.find("a").get("href")
                            )
                    case _: # data columns
                        # cluster of 4 bars, 1 month
                        # one entry per bird/week combo
                        # note: 0-indexed, index = weeknum-1
                        for d in c.find_all("div"):
                            bird_data.append(d.get("class")[0])

                colnum += 1
                # end column processing

            bird["Data"] = bird_data
            # save bird data to table dataset
            results[bird_key]= bird

    return results
def calc_abundance(value):
    result =""

    if 0 <= value <= 1:
        result = "unlikely"
    elif 1 < value <= 4:
        result = "possible"        
    elif 4 < value <= 9:
        result = "likely"        
        
    return result

def calc_species_status(d):
    # pass in monthly abundance data for one species
    # return species status
    #   Breeding
    #   wintering
    #   Migration
    #   Phenology (breeding, wintering, migratory, yearround)
    results = {}
    # initialize values
    # weekdata = np.empty(48, dtype = np.int8)
    weekdata = []
    weeksdetect = []

    #signify weeks to use for each calculation
    breedwks = [21,22,23,24, 25, 26, 27, 28]
    migwks = [13, 14, 15, 16, 36, 37, 38, 39]
    winterwks = [45, 46, 47, 48, 3, 4, 5, 6]
    breedsum = 0
    migsum = 0
    wintersum = 0
    yearsum = 0
    # bdata = []
    # mdata = []
    # wdata = []

    # load data
    week = 1
    for w in d:

        #convert BarSize to number
        bsn = calc_barsizenum(w)
        
        if int(week) in breedwks:
            breedsum += bsn
        if int(week) in migwks:
            migsum += bsn
        if int(week) in winterwks:
            wintersum += bsn

        yearsum += bsn
        weekdata.append(bsn)
        weeksdetect.append(bool(bsn))
        # print(bsn)
        # np.append(weekdata, bsn)
        # print(weekdata)
        week += 1


    # extract data from time periods

    ymean = yearsum / week
    bmean = breedsum / len(breedwks)
    mmean = migsum / len(migwks)
    wmean = wintersum / len(winterwks)

    if sum(weeksdetect) < 7:
        results["Phenology"] = "Occasional"
    elif (mmean*0.5 > wmean and mmean*0.5 > bmean):
        results["Phenology"] = "Migratory"
    elif bmean*0.25 > wmean:
        results["Phenology"] = "Breeding"
    elif (wmean*0.25 > bmean):
        results["Phenology"] = "Wintering"
    else:
        results["Phenology"] = "Yearround"

    results["BreedAbundance"] = calc_abundance(bmean)
    results["MigAbundance"] = calc_abundance(mmean)
    results["WinterAbundance"] = calc_abundance(wmean)
    results["YearAbundance"] = calc_abundance(ymean)
    results["WeeksDetected"] = sum(weeksdetect)
    return results

def remove_chars(t):
    return t.replace("\n", "").replace("\t", "")

def update_trackfile(l):
    trackfile = open("track.json", "w", encoding="utf-8")
    trackfile.write(json.dumps(l))

def main():
    # loop through blocks, hit eBird webpage, and parse data

    # Excel Row counters
    # row counters
    # # rowDA = 2
    rowSP = 2

    # COUNTY COUNTER
    ccount = 0
    for c in counties.keys():

        # do we need to load data from eBird website?        
        loadCounty = False
        currCounty = c
        if currCounty not in scd.keys():
            ## add basic block data
            loadCounty = True

            print(
                nl,
                "=================", nl,
                "retrieving eBird page for", c
            )

            # get html from overview web page
            # # with open("test.html", mode = "r", encoding = "utf-8") as file:
            # #     full_page_html = BeautifulSoup(file, "html.parser")
            # PRODUCTION
            time.sleep(10) #pause for 10 seconds
            base_url = "https://ebird.org/barchart?byr=1900&eyr=2023&bmo=1&emo=12&r="
            r= requests.get(
                base_url + counties[c]
                )
            full_page_html = BeautifulSoup(r.content, "html.parser")
            # END PRODUCTION

            # parse retrieved html - get second table
            for tables in full_page_html.find_all(
                "table",
                attrs = { "class" : "barChart"}
                ):
                if tables.find_all("thead"):
                    content_html = tables
                    break
            # retrieve data from webpage
            # add to species data dict
            scd[c] = get_data(content_html)

            # create blank county record for species list
            csl[c] = []

        else: # county already collected?
            print(nl, "====", nl, c, "already collected", nl, "====")    
        
        #############################################################
        # POPULATE EXCEL
        print("saving data to excel")

        # Loop Species in the County
        for s,v in scd[c].items():
            # loop through species row
            # each species/county/data combo is a row in the table

            # add records to "Data" sheet
            # loop through data items
            # print(s)
            # week = 1
            # # Excel spp data creates too many rows...
            # # for d in v["Data"]:
            # #     wsData[colsData["County"]+str(rowDA)] = currCounty
            # #     wsData[colsData["SpeciesName"]+str(rowDA)] = v["SpeciesName"] 
            # #     wsData[colsData["SpeciesCode"]+str(rowDA)] = v["SpeciesCode"] 
            # #     wsData[colsData["SpeciesCategory"]+str(rowDA)] = v["SpeciesCategory"] 
            # #     wsData[colsData["SpeciesType"]+str(rowDA)] = v["SpeciesType"] 
            # #     wsData[colsData["SpeciesGraphUrl"]+str(rowDA)] = v["SpeciesGraphUrl"] 
            # #     wsData[colsData["Week"]+str(rowDA)] = week 
            # #     wsData[colsData["BarSize"]+str(rowDA)] = d 
            # #     wsData[colsData["BarSizeNum"]+str(rowDA)] = calc_barsizenum(d) 

            # #     # advance spreadsheet row
            # #     rowDA += 1
            
            #add records to "Species" sheet
            if v["SpeciesCategory"] == "species":
                # determine species status
                status = calc_species_status(v["Data"])

                # add data to dict, if recent data loaded from eBird
                if loadCounty:
                    csl[c].append({
                        "SpeciesName" : v["SpeciesName"], 
                        "SpeciesCode" : v["SpeciesCode"], 
                        "SpeciesType" : v["SpeciesType"], 
                        "SpeciesGraphUrl" : v["SpeciesGraphUrl"], 
                        "Phenology" : status["Phenology"], 
                        "BreedAbundance" : status["BreedAbundance"], 
                        "MigAbundance" : status["MigAbundance"], 
                        "WinterAbundance" : status["WinterAbundance"], 
                        "YearAbundance" : status["YearAbundance"], 
                        "WeeksDetected" : status["WeeksDetected"]}
                    )

                # add species to spreadsheet for this county
                wsSpp[colsSpp["County"]+str(rowSP)] = currCounty
                wsSpp[colsSpp["SpeciesName"]+str(rowSP)] = v["SpeciesName"] 
                wsSpp[colsSpp["SpeciesCode"]+str(rowSP)] = v["SpeciesCode"] 
                wsSpp[colsSpp["SpeciesCategory"]+str(rowSP)] = v["SpeciesCategory"] 
                wsSpp[colsSpp["SpeciesType"]+str(rowSP)] = v["SpeciesType"] 
                wsSpp[colsSpp["SpeciesGraphUrl"]+str(rowSP)] = v["SpeciesGraphUrl"]

                wsSpp[colsSpp["Phenology"]+str(rowSP)] = status["Phenology"]
                wsSpp[colsSpp["BreedAbundance"]+str(rowSP)] = status["BreedAbundance"]
                wsSpp[colsSpp["MigAbundance"]+str(rowSP)] = status["MigAbundance"]
                wsSpp[colsSpp["WinterAbundance"]+str(rowSP)] = status["WinterAbundance"]
                wsSpp[colsSpp["YearAbundance"]+str(rowSP)] = status["YearAbundance"]
                wsSpp[colsSpp["WeeksDetected"]+str(rowSP)] = status["WeeksDetected"]

                rowSP += 1
            
            # if rowSP >5:
            #     exit()
            # end species loop
            if loadCounty:
                # newly populated data
                # save data to dictionaries
                sdjson = open("sppdata.json","w")
                sdjson.write(json.dumps(scd))
                sdjson.close()
                
                sljson = open("spplist.json","w")
                sljson.write(json.dumps(csl))
                sljson.close()
            
        print(nl,
            "uploading to atlas cache"
            )
        
        # update mongodb ebird data

        blocksum.update_one(
            {
                "county" : currCounty.upper()
            },
            {
                "$set": {
                    "ebird_county_data": csl[c]
                }
            }
        )

        ccount += 1

        print("atlas cache updated...")
        print(str(ccount), "counties updated.", nl)

            # # TESTING
            # # if ccount > 19: break

    # END COUNTY LOOP

    #########################################################################
    # SAVE WORKBOOK
    # Add tables

    # backup row counters
    # # rowDA -= 1
    rowSP -= 1

    # wsData.add_table(
    #     createTable("sppcountydata", maxDataCol, rowDA)
    # )

    wsSpp.add_table(
        createTable("sppcountylist", maxSppCol, rowSP)
    )

    wb.save(
        os.path.join(
            dirpath,
            update_date + " ebird county stats.xlsx"
        )
    )    

if __name__=="__main__":
    main();